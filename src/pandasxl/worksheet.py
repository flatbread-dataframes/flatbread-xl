# uvmlib/export/worksheet.py
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook

from pandasxl.elements.base import WorksheetElement
from pandasxl.elements.text import TextElement, MultiColumnTextElement
from pandasxl.elements.table import TableElement, GroupedTableElement
from pandasxl.layout import CellPosition


class WorksheetManager:
    """
    Manages the content of a worksheet using element-based rendering.
    Handles creation, positioning, and rendering of various elements.
    """
    def __init__(
        self,
        workbook: Workbook,
        sheet_name: str = "Sheet1",
        show_grid: bool = False,
    ):
        self.workbook = workbook
        self.sheet_name = sheet_name
        self.show_grid = show_grid
        self.elements: list[WorksheetElement] = []
        self._ensure_worksheet()

    def _ensure_worksheet(self):
        """Get or create the worksheet, replacing existing sheet if needed."""
        if self.sheet_name in self.workbook.sheetnames:
            # Remove the existing sheet
            idx = self.workbook.sheetnames.index(self.sheet_name)
            self.workbook.remove(self.workbook[self.sheet_name])
            # Create a new sheet in the same position
            self.worksheet = self.workbook.create_sheet(title=self.sheet_name, index=idx)
        else:
            self.worksheet = self.workbook.create_sheet(title=self.sheet_name)
        self.worksheet.sheet_view.showGridLines = self.show_grid

    def add_title(
        self,
        text: str,
        x_offset: int | None = None,
        y_offset: int | None = None,
    ) -> TextElement:
        """
        Add a worksheet title.

        Parameters
        ----------
        text : str
            Title text
        x_offset : int, optional
            X position (if None, uses the last element's position)
        y_offset : int, optional
            Y position (if None, positions below the last element)

        Returns
        -------
        TextElement
            The created title element
        """
        position = self._get_next_position(x_offset, y_offset)

        element = TextElement(
            text=text,
            x_offset=position.x,
            y_offset=position.y,
            style_preset="title"
        )

        self.elements.append(element)
        element.render(self.worksheet)
        return element

    def _add_regular_table(
        self,
        df: pd.DataFrame,
        x_offset: int,
        y_offset: int,
        title: str | None = None,
        caption: str | None = None,
        **table_kwargs
    ) -> TableElement:
        """Private method to add a regular table."""
        element = TableElement(
            df       = df,
            x_offset = x_offset,
            y_offset = y_offset,
            title    = title,
            caption  = caption,
            **table_kwargs
        )

        self.elements.append(element)
        element.render(self.worksheet)
        return element

    def _add_grouped_table(
        self,
        df: pd.DataFrame,
        x_offset: int | None = None,
        y_offset: int | None = None,
        title: str | None = None,
        caption: str | None = None,
        group_levels: int | list[int] = 0,
        **table_kwargs
    ) -> GroupedTableElement:
        """
        Add a grouped table to the worksheet.

        Parameters
        ----------
        df : pd.DataFrame
            Data to display
        x_offset : int, optional
            X position (if None, uses the last element's position)
        y_offset : int, optional
            Y position (if None, positions below the last element)
        title : str, optional
            Optional title for the table
        caption : str, optional
            Optional caption for the table
        group_levels : int | list[int], default 0
            Index levels to group by
        **table_kwargs
            Additional arguments for the table writer

        Returns
        -------
        GroupedTableElement
            The created grouped table element
        """
        position = self._get_next_position(x_offset, y_offset)

        element = GroupedTableElement(
            df=df,
            x_offset=position.x,
            y_offset=position.y,
            title=title,
            caption=caption,
            group_levels=group_levels,
            **table_kwargs
        )

        self.elements.append(element)
        element.render(self.worksheet)
        return element

    def add_table(
        self,
        df: pd.DataFrame | pd.Series,
        x_offset: int | None = None,
        y_offset: int | None = None,
        title: str | None = None,
        caption: str | None = None,
        group_levels: int | list[int] | None = None,
        **table_kwargs
    ) -> TableElement:
        """
        Add a table to the worksheet. If group_levels is provided, creates a grouped table.

        Parameters
        ----------
        df : pd.DataFrame
            Data to display
        x_offset : int, optional
            X position (if None, uses the last element's position)
        y_offset : int, optional
            Y position (if None, positions below the last element)
        title : str, optional
            Optional title for the table
        caption : str, optional
            Optional caption for the table
        group_levels : int | list[int] | None, default None
            Level(s) to use for grouping. If provided, creates a grouped table.
        **table_kwargs
            Additional arguments for the table writer

        Returns
        -------
        TableElement
            The created table element
        """
        if isinstance(df, pd.Series):
            df = df.to_frame()

        position = self._get_next_position(x_offset, y_offset)

        if group_levels is not None:
            return self._add_grouped_table(
                df           = df,
                x_offset     = position.x,
                y_offset     = position.y,
                title        = title,
                caption      = caption,
                group_levels = group_levels,
                **table_kwargs
            )
        else:
            return self._add_regular_table(
                df       = df,
                x_offset = position.x,
                y_offset = position.y,
                title    = title,
                caption  = caption,
                **table_kwargs
            )

    def add_text(
        self,
        text: str,
        x_offset: int | None = None,
        y_offset: int | None = None,
        style: dict[str, Any] | None = None,
        style_preset: str = "default",
    ) -> TextElement:
        """
        Add a text element to the worksheet.

        Parameters
        ----------
        text : str
            Text content
        x_offset : int, optional
            X position (if None, uses the last element's position)
        y_offset : int, optional
            Y position (if None, positions below the last element)
        style : Dict[str, Any], optional
            Custom style to apply
        style_preset : str, default "default"
            Style preset to use

        Returns
        -------
        TextElement
            The created text element
        """
        position = self._get_next_position(x_offset, y_offset)

        element = TextElement(
            text         = text,
            x_offset     = position.x,
            y_offset     = position.y,
            style        = style,
            style_preset = style_preset
        )

        self.elements.append(element)
        element.render(self.worksheet)
        return element

    def add_multicolumn_text(
        self,
        text: str,
        width: int,
        x_offset: int | None = None,
        y_offset: int | None = None,
        style: dict[str, Any] | None = None,
        style_preset: str = "default",
        row_height: float | None = None,
    ) -> MultiColumnTextElement:
        """
        Add a text element that spans multiple columns with text wrapping.

        Parameters
        ----------
        text : str
            Text content
        width : int
            Number of columns to span
        x_offset : int, optional
            X position (if None, uses the last element's position)
        y_offset : int, optional
            Y position (if None, positions below the last element)
        style : Dict[str, Any], optional
            Custom style to apply
        style_preset : str, default "default"
            Style preset to use
        row_height : float, optional
            Explicit row height in points (if None, Excel will auto-adjust)

        Returns
        -------
        MultiColumnTextElement
            The created multi-column text element
        """
        position = self._get_next_position(x_offset, y_offset)

        element = MultiColumnTextElement(
            text=text,
            width=width,
            x_offset=position.x,
            y_offset=position.y,
            style=style,
            style_preset=style_preset,
            row_height=row_height
        )

        self.elements.append(element)
        element.render(self.worksheet)
        return element

    def autosize_columns(
        self,
        min_width: int = 8,
        max_width: int = 50,
        padding: float = 1.2,
    ):
        """
        Auto-size columns in the worksheet to fit their content.

        Parameters
        ----------
        min_width : int, default 8
            Minimum column width in characters
        max_width : int, default 50
            Maximum column width in characters
        padding : float, default 1.2
            Padding multiplier to add extra space (1.0 = no padding)
        """
        # First pass: identify all cells in merged ranges
        merged_cells = set()
        for merged_range in self.worksheet.merged_cells.ranges:
            # Get all cells in this range
            min_col, min_row, max_col, max_row = merged_range.bounds
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_cells.add((row, col))

        # Track the maximum width needed for each column
        column_widths = {}

        # Only consider table elements for automatic sizing
        table_cells = set()
        for element in self.elements:
            if isinstance(element, TableElement):
                # Get the bounds of the table (not including title/caption)
                min_col = element.table_layout.excel_x_start
                max_col = element.table_layout.excel_x_end
                min_row = element.table_layout.excel_y_start
                max_row = element.table_layout.excel_y_end

                # Add all cells in this table to the set
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        table_cells.add((row, col))

        # Process all actual cells in the worksheet
        for (row, col), cell in self.worksheet._cells.items():
            # Skip cells that are not in tables or are in merged ranges
            if (row, col) not in table_cells or (row, col) in merged_cells:
                continue

            # Get the column index
            col_idx = col

            # Calculate required width based on content
            if cell.value is not None:
                # For string values, use length
                if isinstance(cell.value, str):
                    # Apply font size adjustments if formatted
                    font_multiplier = 1.0
                    if cell.font and cell.font.size:
                        font_multiplier = cell.font.size / 11.0  # Assuming 11 is the base size

                    # Make bold text slightly wider
                    if cell.font and cell.font.bold:
                        font_multiplier *= 1.1

                    content_length = len(cell.value) * font_multiplier
                # For numeric values, use fixed width based on format
                else:
                    content_length = 12  # Default numeric width

                # Update max width for this column if needed
                column_widths[col_idx] = max(
                    column_widths.get(col_idx, min_width),
                    content_length
                )

        # Apply the calculated widths with padding
        for col_idx, width in column_widths.items():
            # Apply padding, but respect min/max constraints
            adjusted_width = min(max(width * padding, min_width), max_width)
            col_letter = self.worksheet.cell(row=1, column=col_idx).column_letter
            self.worksheet.column_dimensions[col_letter].width = adjusted_width

    def _get_next_position(
        self,
        x_offset: int | None = None,
        y_offset: int | None = None
    ) -> CellPosition:
        """
        Calculate the next element position.

        If both x and y are provided, uses those coordinates.
        If only x is provided, positions vertically below the last element.
        If only y is provided, uses that y-coordinate and the x-coordinate of the last element.
        If neither is provided, positions below the last element.

        Returns
        -------
        CellPosition
            The position for the next element
        """
        if len(self.elements) == 0:
            # First element
            return CellPosition(x_offset or 0, y_offset or 0)

        last_element = self.elements[-1]

        if x_offset is not None and y_offset is not None:
            # Both coordinates specified
            return CellPosition(x_offset, y_offset)
        elif x_offset is not None:
            # Only x specified, position below last element with new x
            return CellPosition(x_offset, last_element.y_end + 1)
        elif y_offset is not None:
            # Only y specified, use x from last element
            return CellPosition(last_element.x_start, y_offset)
        else:
            # Neither specified, position below last element
            return last_element.get_position_below()

    @classmethod
    def from_filepath(
        cls,
        filepath: Path | str,
        sheet_name: str = "Sheet",
        show_grid: bool = False,
        open_existing: bool = True,
    ) -> 'WorksheetManager':
        """Create a WorksheetManager associated with a filepath."""
        filepath = Path(filepath) if not isinstance(filepath, Path) else filepath

        if open_existing and filepath.exists():
            wb = load_workbook(filepath)
        else:
            # Create new workbook
            wb = Workbook()

            # Remove the default "Sheet" worksheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # Create manager instance
        manager = cls(
            workbook = wb,
            sheet_name = sheet_name,
            show_grid = show_grid,
        )

        # Store filepath for later saving
        manager.filepath = filepath

        return manager

    def save(self):
        """Save the workbook to the filepath specified during creation."""
        if not hasattr(self, 'filepath'):
            raise ValueError("No filepath associated with this WorksheetManager. Use save_as() instead.")

        self.workbook.save(self.filepath)

    def save_as(self, filepath: Path | str):
        """Save the workbook to a specified filepath."""
        self.workbook.save(filepath)
