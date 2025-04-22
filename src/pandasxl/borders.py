# border.py
from typing import Any

import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side

from pandasxl.layout import TableLayout, CellPosition
from pandasxl.style import BorderStyle


class BorderManager:
    """
    Manages borders in Excel worksheets.
    Handles both structural borders and pattern-based borders.
    """

    def __init__(self, worksheet: Worksheet):
        """
        Initialize with a worksheet.

        Parameters
        ----------
        worksheet : Worksheet
            The openpyxl worksheet to add borders to
        """
        self.worksheet = worksheet

    def add_left_border(
        self,
        cell_position: CellPosition,
        style: BorderStyle,
    ) -> None:
        """Add a left border to a cell."""
        cell = self.worksheet.cell(
            row=cell_position.excel_y,
            column=cell_position.excel_x
        )
        existing_border = cell.border
        side = Side(style=style.value) if style.value else None

        cell.border = Border(
            left=side,
            right=existing_border.right,
            top=existing_border.top,
            bottom=existing_border.bottom
        )

    def add_top_border(
        self,
        cell_position: CellPosition,
        style: BorderStyle,
    ) -> None:
        """Add a top border to a cell."""
        cell = self.worksheet.cell(
            row=cell_position.excel_y,
            column=cell_position.excel_x
        )
        existing_border = cell.border
        side = Side(style=style.value) if style.value else None

        cell.border = Border(
            left=existing_border.left,
            right=existing_border.right,
            top=side,
            bottom=existing_border.bottom
        )

    def add_right_border(
        self,
        cell_position: CellPosition,
        style: BorderStyle,
    ) -> None:
        """Add a right border to a cell."""
        cell = self.worksheet.cell(
            row=cell_position.excel_y,
            column=cell_position.excel_x
        )
        existing_border = cell.border
        side = Side(style=style.value) if style.value else None

        cell.border = Border(
            left=existing_border.left,
            right=side,
            top=existing_border.top,
            bottom=existing_border.bottom
        )

    def add_bottom_border(
        self,
        cell_position: CellPosition,
        style: BorderStyle,
    ) -> None:
        """Add a bottom border to a cell."""
        cell = self.worksheet.cell(
            row=cell_position.excel_y,
            column=cell_position.excel_x
        )
        existing_border = cell.border
        side = Side(style=style.value) if style.value else None

        cell.border = Border(
            left=existing_border.left,
            right=existing_border.right,
            top=existing_border.top,
            bottom=side
        )

    def add_all_borders(
        self,
        cell_position: CellPosition,
        style: BorderStyle,
    ) -> None:
        """Add borders on all sides of a cell."""
        cell = self.worksheet.cell(
            row=cell_position.excel_y,
            column=cell_position.excel_x
        )
        side = Side(style=style.value) if style.value else None

        cell.border = Border(
            left=side,
            right=side,
            top=side,
            bottom=side
        )

    def add_vertical_index_border(
        self,
        layout: TableLayout,
        border_style: BorderStyle = BorderStyle.MEDIUM,
    ) -> None:
        """Add vertical border line between index and data columns."""
        # Add left border to first data column
        data_first_col = layout.data_layout.x_start

        # For each row in the table
        for row_idx in range(layout.y_start, layout.y_end + 1):
            position = CellPosition(data_first_col, row_idx)
            self.add_left_border(position, border_style)

    def add_horizontal_header_border(
        self,
        layout: TableLayout,
        border_style: BorderStyle = BorderStyle.MEDIUM,
    ) -> None:
        """Add horizontal border line between column headers and data rows."""
        # Add top border to first data row
        data_first_row = layout.data_layout.y_start

        # For each column in the table
        for col_idx in range(layout.x_start, layout.x_end + 1):
            position = CellPosition(col_idx, data_first_row)
            self.add_top_border(position, border_style)

    def add_level_borders(
        self,
        layout: TableLayout,
        row_spans: list[list[dict[str, Any]]],
        column_spans: list[list[dict[str, Any]]],
        min_border_level: int = 1,
        border_style: BorderStyle = BorderStyle.THIN,
    ) -> None:
        """Add borders between different levels of MultiIndex.

        Parameters
        ----------
        layout : TableLayout
            Table layout information
        row_spans : list[list[dict[str, Any]]]
            Spans information for rows from spans.get_level_spans()
        column_spans : list[list[dict[str, Any]]]
            Spans information for columns from spans.get_level_spans()
        min_border_level : int, default 1
            Minimum level to add borders for
        border_style : BorderStyle, default BorderStyle.THIN
            Style to use for the borders
        """
        # Add row level borders if we have spans above min_border_level
        if len(row_spans) > min_border_level:
            self._add_row_level_borders(
                layout,
                row_spans,
                min_border_level,
                border_style,
            )

        # Add column level borders if we have spans above min_border_level
        if len(column_spans) > min_border_level:
            self._add_column_level_borders(
                layout,
                column_spans,
                min_border_level,
                border_style,
            )

    # For _add_row_level_borders in borders.py
    def _add_row_level_borders(
        self,
        layout: TableLayout,
        row_spans: list[list[dict[str, Any]]],
        min_border_level: int,
        border_style: BorderStyle,
    ) -> None:
        """Add horizontal borders between different levels in row MultiIndex.

        Parameters
        ----------
        layout : TableLayout
            Table layout information
        row_spans : list[list[dict[str, Any]]]
            Spans information for rows
        min_border_level : int
            Minimum number of levels to exclude from the end
        border_style : BorderStyle
            Style to use for the borders
        """
        # Calculate the maximum level to show borders for
        # If min_border_level is 1 and we have 3 levels (0, 1, 2), we draw borders for levels 0 and 1
        max_level_to_show = len(row_spans) - min_border_level

        # Iterate through levels we want to show borders for
        for level in range(max_level_to_show):
            spans = row_spans[level]

            # For each span at this level
            for span in spans:
                # Skip first span (no border needed at the start)
                if span['start'] > 0:
                    # Add a horizontal border at the start of each span
                    row_pos = layout.index.y_start + span['start']

                    # Add border across the entire table width
                    for x in range(layout.x_start, layout.x_end + 1):
                        position = CellPosition(x, row_pos)
                        self.add_top_border(position, border_style)

    # For _add_column_level_borders in borders.py
    def _add_column_level_borders(
        self,
        layout: TableLayout,
        column_spans: list[list[dict[str, Any]]],
        min_border_level: int,
        border_style: BorderStyle,
    ) -> None:
        """Add vertical borders between different levels in column MultiIndex.

        Parameters
        ----------
        layout : TableLayout
            Table layout information
        column_spans : list[list[dict[str, Any]]]
            Spans information for columns
        min_border_level : int
            Minimum number of levels to exclude from the end
        border_style : BorderStyle
            Style to use for the borders
        """
        # Calculate the maximum level to show borders for
        # If min_border_level is 1 and we have 3 levels (0, 1, 2), we draw borders for levels 0 and 1
        max_level_to_show = len(column_spans) - min_border_level

        # Iterate through levels we want to show borders for
        for level in range(max_level_to_show):
            spans = column_spans[level]

            # For each span at this level
            for span in spans:
                # Skip first span (no border needed at the start)
                if span['start'] > 0:
                    # Add a vertical border at the start of each span
                    col_pos = layout.columns.x_start + span['start']

                    # Add border down the entire table height
                    for y in range(layout.y_start, layout.y_end + 1):
                        position = CellPosition(col_pos, y)
                        self.add_left_border(position, border_style)

    def add_custom_borders(
        self,
        layout: TableLayout,
        row_borders: list[bool],
        column_borders: list[bool],
        border_style: BorderStyle = BorderStyle.THIN
    ) -> None:
        """
        Add borders at positions specified by pre-processed border lists.

        Parameters
        ----------
        layout : TableLayout
            Table layout information
        row_borders : list[bool]
            List where index i is True if row i needs a border
        column_borders : list[bool]
            List where index j is True if column j needs a border
        border_style : BorderStyle
            Style to use for the borders
        """
        # Apply row borders where marked
        for i, needs_border in enumerate(row_borders):
            if needs_border:
                row_pos = layout.index.y_start + i
                for x in range(layout.x_start, layout.x_end + 1):
                    position = CellPosition(x, row_pos)
                    self.add_top_border(position, border_style)

        # Apply column borders where marked
        for j, needs_border in enumerate(column_borders):
            if needs_border:
                col_pos = layout.columns.x_start + j
                for y in range(layout.y_start, layout.y_end + 1):
                    position = CellPosition(col_pos, y)
                    self.add_left_border(position, border_style)
